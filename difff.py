from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from tkinter import Tk, filedialog
import difflib
import os

root = Tk()
root.withdraw()  # Tkinterのメインウィンドウを非表示にする

# 差分の書式指定
formatting = {"color": "FF0000", "b": True, "sz": 14} # ①
red_large_font = InlineFont(**formatting)

# 比較の最初の行を設定
comparison_start_row = 2 # ②

# エクセルファイルを選択
file = filedialog.askopenfilename(title='エクセルファイルを選択してください', filetypes=[("EXCEL", "*.xlsx")]) # ③

# エクセルファイルを開く
wb = load_workbook(file)
ws = wb.active

for row in range(comparison_start_row, ws.max_row + 1):
    # 比較対象のセルを取得
    before_cell = ws[f'A{row}'] # ④
    after_cell = ws[f'B{row}']
    word1 = ""
    word2 = ""
    
    if before_cell.value is None and after_cell.value is None:
        # セルの値が空の場合は何もしない
        pass
    elif before_cell.value is None:
        # セルの値が空の場合は、もう片方のセルを赤字にする
        after_cell.font = Font(**formatting)
    elif after_cell.value is None:
        # セルの値が空の場合は、もう片方のセルを赤字にする
        before_cell.font = Font(**formatting)
    elif before_cell.value != after_cell.value:
        # セルの値が異なる場合は、差分を取得する
        ndiff = difflib.ndiff(str(before_cell.value), str(after_cell.value)) # ⑤

        # 差分の箇所に$を付けて、視覚化する
        for i in ndiff:
            print(i)
            if i.startswith(" "):
                i = i[2:].replace(" ", "\u00A0")  # 半角の空白をノーブレークスペースに置き換え
                word1 += i
                word2 += i
            elif i.startswith("- "):
                i = i[2:].replace(" ", "\u00A0")  # 半角の空白をノーブレークスペースに置き換え
                word1 += f"${i}"
            elif i.startswith("+ "):
                i = i[2:].replace(" ", "\u00A0")  # 半角の空白をノーブレークスペースに置き換え
                word2 += f"${i}"

        rich_text_before = []
        i = 0

        while i < len(word1):
            if word1[i] == "$":
                # $がある場合は、差分の書式を適用する
                i += 1
                rich_text_before.append(TextBlock(red_large_font, word1[i]))
            else:
                rich_text_before.append(word1[i])
            i += 1
        before_cell.value = CellRichText(*rich_text_before)

        rich_text_after = []
        i = 0
        while i < len(word2):
            if word2[i] == "$":
                # $がある場合は、差分の書式を適用する
                i += 1
                rich_text_after.append(TextBlock(red_large_font, word2[i]))
            else:
                rich_text_after.append(word2[i])
            i += 1
        after_cell.value = CellRichText(*rich_text_after)

wb.save(file)
os.system(f'open "{file}"')